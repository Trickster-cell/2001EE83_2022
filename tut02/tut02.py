from openpyxl import Workbook
from openpyxl import load_workbook

from datetime import datetime
start_time = datetime.now()


def decide_octant(u, v, w):
    # Function to decide the octant on the basis of given values of u, v, and w
    if(u > 0 and v > 0 and w > 0):
        return "+1"
    elif(u > 0 and v > 0 and w < 0):
        return "-1"
    elif(u < 0 and v < 0 and w > 0):
        return "+3"
    elif(u < 0 and v < 0 and w < 0):
        return "-3"
    elif(u < 0 and v > 0 and w > 0):
        return "+2"
    elif(u < 0 and v > 0 and w < 0):
        return "-2"
    elif(u > 0 and v < 0 and w > 0):
        return "+4"
    elif(u > 0 and v < 0 and w < 0):
        return "-4"


def mainfun(Mod=5000):
    # Main function which we will call at the bottom. The Mod value can be changed here in function parameter, or when the function is called.
    wb = load_workbook(r'input_octant_transition_identify.xlsx')

    sheet = wb.active
    # reading the input file

    cnt = 0
    arr = [0.0, 0.0, 0.0]
    # array to add all the u,v,w values for getting average and etc.

    headings = ["U Avg", "V Avg", "W Avg", "U'=U - U avg", "V'=V - V avg",
                "W'=W - W avg", "Octant", " ", " ", "+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]

    # list that will store all the heading for each column

    modstr = "Mod " + (str)(Mod)
# modstr is basically a string like Mod 5000

    row_count = sheet.max_row
    column_count = sheet.max_column
    # print(row_count, column_count)

    for col in range(5, 5+len(headings)):
        sheet.cell(row=1, column=col).value = headings[col-5]
# printed remaining headings that weren't in the input file

    for i in range(2, row_count + 1):
        for j in range(2, column_count + 1):
            data = sheet.cell(row=i, column=j).value
            x = float(data)
            arr[j-2] += x
    for i in range(3):
        arr[i] /= row_count-1
    # print(arr)
    # calculated the average values

    for i in range(5, 8):
        sheet.cell(row=2, column=i).value = arr[i-5]
    for i in range(2, row_count+1):
        for j in range(8, 11):
            sheet.cell(row=i, column=j).value = sheet.cell(
                row=i, column=j-6).value - arr[j-8]
# filled the columns in with u', v', and w'

    dict = {"+1": 0, "-1": 0, "+2": 0, "-2": 0,
            "+3": 0, "-3": 0, "+4": 0, "-4": 0}
# made a dictionary to store the overall count of each of the octants

    keys = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
# list to store all the keys of octants

    for i in range(2, row_count+1):
        u = sheet.cell(row=i, column=8).value
        v = sheet.cell(row=i, column=9).value
        w = sheet.cell(row=i, column=10).value
        sheet.cell(row=i, column=11).value = (str)(decide_octant(u, v, w))
        dict[decide_octant(u, v, w)] += 1
    # filled the octant column with the help of octant decide function

    sheet.cell(row=2, column=13).value = "Overall Count"
    sheet.cell(row=3, column=12).value = "User Input"
    sheet.cell(row=3, column=13).value = modstr
# wrote the mentioned things in the file

    for i in range(14, 22):
        sheet.cell(row=2, column=i).value = dict[keys[i-14]]
# filled the overall count table

    modcnt = row_count//Mod
    modcnt += 1
    if row_count % Mod == 0:
        modcnt -= 1
    # print(modcnt)

    dict2 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3,
             "+3": 4, "-3": 5, "+4": 6, "-4": 7}
#  dictionary to store the offset values corresponding to the octants
#  this will help us while we will be filling mod count octant transitions

    arr2 = [0]*8
    finarr = []
    for i in range(modcnt):
        finarr.append(arr2.copy())

# made an empty 2d array to store the values for table

    ind = 0
    cnt = 0
    for i in range(2, row_count+1):
        oct = sheet.cell(row=i, column=11).value
        j = dict2[oct]
        finarr[ind][j] += 1
        cnt += 1
        if cnt == Mod:
            ind += 1
            cnt = 0

# stored the values for the overall octant count table

    lv = 0
    hv = Mod-1

    for i in range(4, 4+modcnt):
        sheet.cell(row=i, column=13).value = str(lv) + " - " + str(hv)
        lv = hv+1
        hv += Mod
        hv = min(hv, row_count)
# filled the octant table first column i.e. 0-4999

    for i in range(4, 4+modcnt):
        for j in range(14, 22):
            sheet.cell(row=i, column=j).value = finarr[i-4][j-14]

    sheet.cell(row=4+modcnt, column=13).value = "Verified"
    for i in range(14, 22):
        sheet.cell(row=4+modcnt, column=i).value = 0
        for j in range(4, 4+modcnt):
            sheet.cell(
                row=4+modcnt, column=i).value += sheet.cell(row=j, column=i).value
# filled the overall octant count table

    lastval = 4+modcnt

    lastval += 3
    sheet.cell(row=lastval, column=13).value = "Overall Transition Count"
    sheet.cell(row=lastval+1, column=14).value = "To"
    sheet.cell(row=lastval+3, column=12).value = "From"
    sheet.cell(row=lastval+2, column=13).value = "Count"
# filled the headings for the overall transition count table

    for i in range(8):
        sheet.cell(row=lastval+2, column=14+i).value = keys[i]
    for i in range(8):
        sheet.cell(row=lastval+3+i, column=13).value = keys[i]
# filled the titles for rows and columns in transition count table
    lv = 0
    hv = Mod-1

    constr = lastval+3

    strtvals = []

    for i in range(modcnt):
        lastval += 14
        sheet.cell(row=lastval+1, column=13).value = str(lv)+" - "+str(hv)
        sheet.cell(row=lastval, column=13).value = "Mod Transition Count"
        sheet.cell(row=lastval+1, column=14).value = "To"
        sheet.cell(row=lastval+2, column=13).value = "Count"
        sheet.cell(row=lastval+3, column=12).value = "From"
        strtvals.append(lastval+3)
        lv = hv+1
        hv += Mod
        hv = min(hv, row_count)

        for i in range(8):
            sheet.cell(row=lastval+2, column=14+i).value = keys[i]
        for i in range(8):
            sheet.cell(row=lastval+3+i, column=13).value = keys[i]

    # filled the octant table headings

    # print(finarr)

    k = 0
    cnt = 0
    prev = sheet.cell(row=2, column=11).value
    # print(strtvals)
    # print(prev)
    for i in range(3, row_count+1):
        curr = sheet.cell(row=i, column=11).value
        tempprev = dict2[prev]
        tempcurr = dict2[curr]
        temp1 = sheet.cell(row=constr+tempprev, column=14+tempcurr).value
        if(temp1 != None):
            sheet.cell(row=constr+tempprev, column=14 +
                       tempcurr).value = int(temp1)+1
        else:
            sheet.cell(row=constr+tempprev, column=14+tempcurr).value = 1

        temp = sheet.cell(row=strtvals[k]+tempprev, column=14+tempcurr).value
        if(temp != None):
            sheet.cell(row=strtvals[k]+tempprev,
                       column=14+tempcurr).value = int(temp)+1
        else:
            sheet.cell(row=strtvals[k]+tempprev, column=14+tempcurr).value = 1

        cnt += 1
        prev = str(curr)
        if cnt % Mod == 0:
            k += 1
# finally filled the octant tables for different mod ranges

    for val in strtvals:
        for i in range(8):
            for j in range(8):
                if(sheet.cell(row=val+i, column=14+j).value == None):
                    sheet.cell(row=val+i, column=14+j).value = 0

# and then finally set the empty cells in the octant tables for different mod ranges to 0

    wb.save("output_octant_transition_identify.xlsx")
# saved the sheet in output file
    print("Program finally executed with value of Mod =", Mod)
# printing the code success message


mainfun(5000)
#  calling the main function
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
