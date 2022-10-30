from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
start_time = datetime.now()

def decide_octant(u, v, w):
    # Function to decide the octant on the basis of given values of u, v, and w
    if(u > 0 and v > 0 and w > 0):
        return "+1"
    elif(u > 0 and v > 0 and w < 0):
        return "-1"
    elif(u < 0 and v < 0 and w > 0):
        return "+3"
    elif(u < 0 and v < 0 and w < 0):
        return "-3"
    elif(u < 0 and v > 0 and w > 0):
        return "+2"
    elif(u < 0 and v > 0 and w < 0):
        return "-2"
    elif(u > 0 and v < 0 and w > 0):
        return "+4"
    elif(u > 0 and v < 0 and w < 0):
        return "-4"


# Main function which we will call at the bottom.
def fun():
    wb = load_workbook(r'input_octant_longest_subsequence.xlsx')

    sheet = wb.active
    # reading the input file

    cnt = 0
    arr = [0.0, 0.0, 0.0]
    # array to add all the u,v,w values for getting average and etc.

    headings = ["U Avg", "V Avg", "W Avg", "U'=U - U avg", "V'=V - V avg",
                "W'=W - W avg", "Octant", " ", "Count", "Longest Subsequence Count", "Count"]

    # list that will store all the heading for each column

    row_count = sheet.max_row
    column_count = sheet.max_column
    # print(row_count, column_count)

    for col in range(5, 5+len(headings)):
        sheet.cell(row=1, column=col).value = headings[col-5]
    # printed remaining headings that weren't in the input file

    for i in range(2, row_count + 1):
        for j in range(2, 4 + 1):
            data = sheet.cell(row=i, column=j).value
            x = float(data)
            arr[j-2] += x
    for i in range(3):
        arr[i] /= row_count-1
    # print(arr)
    # calculated the average values

    for i in range(5, 8):
        sheet.cell(row=2, column=i).value = arr[i-5]
    for i in range(2, row_count+1):
        for j in range(8, 11):
            sheet.cell(row=i, column=j).value = sheet.cell(
                row=i, column=j-6).value - arr[j-8]
        # filled the columns in with u', v', and w'

    dict = {"+1": 0, "-1": 0, "+2": 0, "-2": 0,
            "+3": 0, "-3": 0, "+4": 0, "-4": 0}
    # made a dictionary to store the maximum length of the continuous subsequence

    keys = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    # list to store all the keys of octants

    for i in range(2, row_count+1):
        u = sheet.cell(row=i, column=8).value
        v = sheet.cell(row=i, column=9).value
        w = sheet.cell(row=i, column=10).value
        sheet.cell(row=i, column=11).value = (str)(decide_octant(u, v, w))
        # dict[decide_octant(u, v, w)] += 1

    for i in range(2, 10):
        sheet.cell(row=i, column=13).value = str(keys[i-2])
    # filled the octant column with the help of octant decide function

    prev = str("+1")
    prcnt = 0
    for i in range(2, row_count+1):
        k = str(sheet.cell(row=i, column=11).value)
        # print(k,prev)
        if k is not prev:
            # print(k)
            temp = max(dict[prev], prcnt)
            dict[prev] = temp
            prcnt = 1
            prev = k
        else:
            prcnt += 1
        dict[prev] = max(dict[prev], prcnt)
    # filled the dictionary to store the maximum lengths
    # print(dict)

    for i in range(8):
        sheet.cell(row=2+i, column=14).value = dict[keys[i]]

    dict2 = {"+1": 0, "-1": 0, "+2": 0, "-2": 0,
             "+3": 0, "-3": 0, "+4": 0, "-4": 0}

    prev = "+1"
    prcnt = 0
    for i in range(2, row_count+1):
        k = str(sheet.cell(row=i, column=11).value)
        # print(k,prev)
        if k is not prev:
            # print(k)
            if prcnt == dict[prev]:
                dict2[prev] += 1
            prcnt = 1
            prev = k
        else:
            prcnt += 1
    # print the values of maximum occurence of the continuous octants
    # print(dict2)

    for i in range(8):
        sheet.cell(row=2+i, column=15).value = dict2[keys[i]]

    for idx, col in enumerate(sheet.columns, 1):
        sheet.column_dimensions[get_column_letter(idx)].auto_size = True
    # resized the columns according to their content

    wb.save("output_octant_longest_subsequence.xlsx")
# saved the sheet in output file
fun()
print("Program finally executed.")
# printing the code success message
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
