from ast import Str
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def decide_octant(u, v, w):
    # Function to decide the octant on the basis of given values of u, v, and w
    if(u > 0 and v > 0 and w > 0):
        return "+1"
    elif(u > 0 and v > 0 and w < 0):
        return "-1"
    elif(u < 0 and v < 0 and w > 0):
        return "+3"
    elif(u < 0 and v < 0 and w < 0):
        return "-3"
    elif(u < 0 and v > 0 and w > 0):
        return "+2"
    elif(u < 0 and v > 0 and w < 0):
        return "-2"
    elif(u > 0 and v < 0 and w > 0):
        return "+4"
    elif(u > 0 and v < 0 and w < 0):
        return "-4"


# Main function which we will call at the bottom. The Mod value can be changed here in function parameter, or when the function is called.
wb = load_workbook(r'input_octant_longest_subsequence.xlsx')

sheet = wb.active
# reading the input file

cnt = 0
arr = [0.0, 0.0, 0.0]
# array to add all the u,v,w values for getting average and etc.

headings = ["U Avg", "V Avg", "W Avg", "U'=U - U avg", "V'=V - V avg",
            "W'=W - W avg", "Octant", " ", "Octant", "Longest Subsequence Count", "Count", " ", "Octant", "Longest Subsequence Count", "Count"]

# list that will store all the heading for each column

# modstr = "Mod " + (str)(Mod)
# modstr is basically a string like Mod 5000

row_count = sheet.max_row
column_count = sheet.max_column
   # print(row_count, column_count)

for col in range(5, 5+len(headings)):
    sheet.cell(row=1, column=col).value = headings[col-5]
# printed remaining headings that weren't in the input file

for i in range(2, row_count + 1):
    for j in range(2, 4 + 1):
        data = sheet.cell(row=i, column=j).value
        x = float(data)
        arr[j-2] += x
for i in range(3):
    arr[i] /= row_count-1
# print(arr)
# calculated the average values

for i in range(5, 8):
    sheet.cell(row=2, column=i).value = arr[i-5]
for i in range(2, row_count+1):
    for j in range(8, 11):
        sheet.cell(row=i, column=j).value = sheet.cell(
            row=i, column=j-6).value - arr[j-8]
# filled the columns in with u', v', and w'

dict = {"+1": 0, "-1": 0, "+2": 0, "-2": 0,"+3": 0, "-3": 0, "+4": 0, "-4": 0}
# made a dictionary to store the overall count of each of the octants

keys = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
# list to store all the keys of octants

for i in range(2, row_count+1):
    u = sheet.cell(row=i, column=8).value
    v = sheet.cell(row=i, column=9).value
    w = sheet.cell(row=i, column=10).value
    sheet.cell(row=i, column=11).value = (str)(decide_octant(u, v, w))
    # dict[decide_octant(u, v, w)] += 1
    
for i in range (2, 10):
    sheet.cell(row=i, column=13).value = str(keys[i-2])
# filled the octant column with the help of octant decide function

prev = str("+1")
prcnt = 0
for i in range (2, row_count+1):
    k = str(sheet.cell(row=i, column=11).value)
    # print(k,prev)
    if k is not prev:
        # print(k)
        temp=max(dict[prev], prcnt)
        dict[prev] = temp
        prcnt = 1
        prev = k
    else:
        prcnt+=1
dict[prev] = max(dict[prev], prcnt)     

# print(dict)


for i in range (8):
    sheet.cell(row=2+i, column=14).value = dict[keys[i]]

frq_of_maxm_subs = {"+1": 0, "-1": 0, "+2": 0, "-2": 0,"+3": 0, "-3": 0, "+4": 0, "-4": 0}
#dictionary to store the frequency of maximum length of subsequence 

starting_points = {"+1": [], "-1": [], "+2": [], "-2": [],"+3": [], "-3": [], "+4": [], "-4": []}
# dictionary to store the starting point times of maximum length subsequence
ending_points = {"+1": [], "-1": [], "+2": [], "-2": [],"+3": [], "-3": [], "+4": [], "-4": []}
# dictionary to store the ending point times of maximum length subsequence


prev = "+1"
prcnt = 0
stp = sheet.cell(row=2, column=1).value
etp = sheet.cell(row=2, column=1).value
for i in range(2, row_count+1):
    k = str(sheet.cell(row=i, column=11).value)
    # print(k,prev)
    if k is not prev:
        # print(k)
        if prcnt == dict[prev]:
            frq_of_maxm_subs[prev]+=1
            # increasing the count of maximum length subsequence
            starting_points[prev].append(stp)
            ending_points[prev].append(etp)
            # storing the starting and ending point of the subsequences
        prcnt = 1
        prev = k
        stp = sheet.cell(row = i, column=1).value
    else:
        prcnt+=1
    etp=sheet.cell(row=i,column=1).value
# print(frq_of_maxm_subs)

# print(starting_points)
# print(ending_points)

# printing the subsequence time range values.
k = 2

for i in range(8):
    sheet.cell(row=k, column=17).value =  keys[i]
    sheet.cell(row=k, column=18).value =  dict[keys[i]]
    sheet.cell(row=k, column=19).value =  frq_of_maxm_subs[keys[i]]
    k+=1
    sheet.cell(row=k, column=17).value =  "Time"
    sheet.cell(row=k, column=18).value =  "From"
    sheet.cell(row=k, column=19).value =  "To"
    k+=1
    for j in range (frq_of_maxm_subs[keys[i]]):
        sheet.cell(row=k, column=18).value = starting_points[keys[i]][j]
        sheet.cell(row=k, column=19).value = ending_points[keys[i]][j]
        k+=1
    

for i in range (8):
    sheet.cell(row=2+i, column=15).value = frq_of_maxm_subs[keys[i]]
    


for idx, col in enumerate(sheet.columns, 1):
    sheet.column_dimensions[get_column_letter(idx)].auto_size = True
# adjusting the column width according to the contents.

wb.save("output_octant_longest_subsequence.xlsx")
# saved the sheet in output file
print("Program finally executed.")
# printing the code success message


