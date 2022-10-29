from datetime import datetime
start_time = datetime.now()

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# import numpy as np


def decide_octant(u, v, w):
    # Function to decide the octant on the basis of given values of u, v, and w
    if (u > 0 and v > 0 and w > 0):
        return "1"
    elif (u > 0 and v > 0 and w < 0):
        return "-1"
    elif (u < 0 and v < 0 and w > 0):
        return "3"
    elif (u < 0 and v < 0 and w < 0):
        return "-3"
    elif (u < 0 and v > 0 and w > 0):
        return "2"
    elif (u < 0 and v > 0 and w < 0):
        return "-2"
    elif (u > 0 and v < 0 and w > 0):
        return "4"
    elif (u > 0 and v < 0 and w < 0):
        return "-4"


def FunToSort(rowi, sheet, list1, keys, keynames, dict4first):
    # function to print the sorted index on the sheet 
    list = []
    for i in range(8):
        list.append([sheet.cell(row=rowi, column=14+i).value, i+1])
    list.sort(reverse=True)
    # appended the values on a list and then reverse sorted it
    for i in range(8):
        # print(list[i][1])
        list1.append(list[i][1])
    for i in range(8):
        sheet.cell(row=rowi, column=21+list1[i]).value = i+1
    sheet.cell(row=rowi, column=30).value = keys[list1[0]-1]
    dict4first[keys[list1[0]-1]]+=1
    sheet.cell(row=rowi, column=31).value = keynames[list1[0]-1]
    # finally printed the values on the sheet


def mainfun(Mod=5000):
    # Main function which we will call at the bottom. The Mod value can be changed here in function parameter, or when the function is called.
    wb = load_workbook(r'octant_input.xlsx')

    sheet = wb.active
    # reading the input file

    cnt = 0
    arr = [0.0, 0.0, 0.0]
    # array to add all the u,v,w values for getting average and etc.

    headings = ["U Avg", "V Avg", "W Avg", "U'=U - U avg", "V'=V - V avg",
                "W'=W - W avg", "Octant", " ", " ", "1", "-1", "2", "-2", "3", "-3", "4", "-4"]

    # list that will store all the heading for each column

    modstr = "Mod " + (str)(Mod)
# modstr is basically a string like Mod 5000

    row_count = sheet.max_row
    column_count = sheet.max_column
    # print(row_count, column_count)

    for col in range(5, 5+len(headings)):
        sheet.cell(row=1, column=col).value = headings[col-5]
# printed remaining headings that weren't in the input file

    for i in range(2, row_count + 1):
        for j in range(2, column_count + 1):
            data = sheet.cell(row=i, column=j).value
            x = float(data)
            arr[j-2] += x
    for i in range(3):
        arr[i] /= row_count-1
    # print(arr)
    # calculated the average values

    for i in range(5, 8):
        sheet.cell(row=2, column=i).value = arr[i-5]
    for i in range(2, row_count+1):
        for j in range(8, 11):
            sheet.cell(row=i, column=j).value = sheet.cell(
                row=i, column=j-6).value - arr[j-8]
# filled the columns in with u', v', and w'

    dict = {"1": 0, "-1": 0, "2": 0, "-2": 0,
            "3": 0, "-3": 0, "4": 0, "-4": 0}
# made a dictionary to store the overall count of each of the octants

    keys = ["1", "-1", "2", "-2", "3", "-3", "4", "-4"]
    keynames = ["Internal outward interaction",
                "External outward interaction",
                "External Ejection",
                "Internal Ejection",
                "External inward interaction",
                "Internal inward interaction",
                "Internal sweep",
                "External sweep"
                ]
# lists to store all the keys of octants and their names

    for i in range(2, row_count+1):
        u = sheet.cell(row=i, column=8).value
        v = sheet.cell(row=i, column=9).value
        w = sheet.cell(row=i, column=10).value
        sheet.cell(row=i, column=11).value = (str)(decide_octant(u, v, w))
        dict[decide_octant(u, v, w)] += 1
    # filled the octant column with the help of octant decide function

    sheet.cell(row=2, column=13).value = "Overall Count"
    sheet.cell(row=3, column=12).value = "User Input"
    sheet.cell(row=3, column=13).value = modstr
# wrote the mentioned things in the file

    for i in range(14, 22):
        sheet.cell(row=2, column=i).value = dict[keys[i-14]]
# filled the overall count table

    modcnt = row_count//Mod
    modcnt += 1
    if row_count % Mod == 0:
        modcnt -= 1
    # print(modcnt)

    dict2 = {"1": 0, "-1": 1, "2": 2, "-2": 3,
             "3": 4, "-3": 5, "4": 6, "-4": 7}
#  dictionary to store the offset values corresponding to the octants
#  this will help us while we will be filling mod count octant transitions

    arr2 = [0]*8
    finarr = []
    for i in range(modcnt):
        finarr.append(arr2.copy())

# made an empty 2d array to store the values for table

    ind = 0
    cnt = 0
    for i in range(2, row_count+1):
        oct = sheet.cell(row=i, column=11).value
        j = dict2[oct]
        finarr[ind][j] += 1
        cnt += 1
        if cnt == Mod:
            ind += 1
            cnt = 0

# stored the values for the overall octant count table

    lv = 0
    hv = Mod-1

    for i in range(4, 4+modcnt):
        sheet.cell(row=i, column=13).value = str(lv) + " - " + str(hv)
        lv = hv+1
        hv += Mod
        hv = min(hv, row_count)
# filled the octant table first column i.e. 0-4999

    for i in range(4, 4+modcnt):
        for j in range(14, 22):
            sheet.cell(row=i, column=j).value = finarr[i-4][j-14]

    # sheet.cell(row=4+modcnt, column=13).value = "Verified"
# filled the overall octant count table


# and then finally set the empty cells in the octant tables for different mod ranges to 0

    sheet.insert_rows(1)
    # inserted a row at top
    for i in range(len(keys)):
        sheet.cell(row=1, column=22+i).value = keys[i]
        sheet.cell(row=2, column=22+i).value = "Rank" + str(1+i)
    # headings in row1
    sheet.cell(row=2, column=22+8).value = "Rank1 Octant ID"
    sheet.cell(row=2, column=22+9).value = "Rank1 Octant Name"

    listtemp = []
    dicttrash = {"1": 0, "-1": 1, "2": 2, "-2": 3,
             "3": 4, "-3": 5, "4": 6, "-4": 7}
    # just a trash dictionary to make the function work
    FunToSort(3, sheet, listtemp, keys, keynames, dicttrash)
    # called the function to write sorted values on the sheet
    dicttrash.clear
    # print(listtemp)
    
    k = row_count//Mod
    sz = k+1;
    if(row_count%Mod==0):
        sz-=1
    # calculated number of rows for the table
    dict5 = {"1": 0, "-1": 0, "2": 0, "-2": 0,
            "3": 0, "-3": 0, "4": 0, "-4": 0}
    # dictionary to make the function work and will store how many times
    # a particular octant was the most occuring octant
    for i in range(sz):
        listtemp2 = []
        FunToSort(5+i, sheet, listtemp2, keys, keynames, dict5)
        listtemp2.clear
    # implemented function for all the rows starting from 5..
    
    # print(dict5)
    endoftab = 4+sz
    # calculate the row where the table ends
    endoftab+=4
    # giving a proper gap between tables
    
    
     
    # now filling the table for frequency of rank1 octants
    sheet.cell(row=14, column=endoftab).value = "Octant ID"
    sheet.cell(row=15, column=endoftab).value = "Octant Name"
    sheet.cell(row=16, column=endoftab).value = "Count of Rank 1 Mod Values"
    endoftab+=1
    for i in range(8):
        sheet.cell(row = endoftab+i, column=14).value = keys[i]
        sheet.cell(row = endoftab+i, column=15).value = keynames[i]
        sheet.cell(row = endoftab+i, column=16).value = dict5[keys[i]]
        
    
    
    
    for idx, col in enumerate(sheet.columns, 1):
        sheet.column_dimensions[get_column_letter(idx)].auto_size = True
    # adjusted the columns size according to their contents
    
    
    
    wb.save("octant_output_ranking_excel.xlsx")
    # saved the sheet in output file
    print("Program finally executed with value of Mod =", Mod)
    # printing the code success message


mainfun(5000)
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
#  calling the main function
