import openpyxl
from openpyxl import load_workbook
import pandas as pd
import os
import datetime


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# imported required library


def send_mail(fromaddr, frompasswd, toaddr, msg_subject, msg_body, file_path):
# Mail sending function
    try:
        msg = MIMEMultipart()
        print("[+] Message Object Created")
    except:
        print("[-] Error in Creating Message Object")
        return

    msg['From'] = fromaddr

    msg['To'] = toaddr

    msg['Subject'] = msg_subject

    body = msg_body

    msg.attach(MIMEText(body, 'plain'))

    filename = file_path
    attachment = open(filename, "rb")

    p = MIMEBase('application', 'octet-stream')

    p.set_payload((attachment).read())

    encoders.encode_base64(p)

    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)

    try:
        msg.attach(p)
        print("[+] File Attached")
    except:
        print("[-] Error in Attaching file")
        return

    try:
        # s = smtplib.SMTP('smtp.gmail.com', 587) if you are sending mail from Gmail, for Gmail use app password in password
        # s = smtplib.SMTP('mail.iitp.ac.in', 587) if you are sending mail from zimbra mail
        s = smtplib.SMTP('stud.iitp.ac.in', 587) #if you are sending mail from zimbra stud
        print("[+] SMTP Session Created")
    except:
        print("[-] Error in creating SMTP session")
        return

    s.starttls()

    try:
        s.login(fromaddr, frompasswd)
        print("[+] Login Successful")
    except:
        print("[-] Login Failed")

    text = msg.as_string()

    try:
        s.sendmail(fromaddr, toaddr, text)
        print("[+] Mail Sent successfully")
    except:
        print('[-] Mail not sent')

    s.quit()


def isEmail(x):
    if ('@' in x) and ('.' in x):
        return True
    else:
        return False




FROM_ADDR = "yash_2001ee83@iitp.ac.in" #change sender's mail id
FROM_PASSWD = "CHANGE IT" #Password
TO_ADDR = "cs3842022@gmail.com"

Subject = "Consolidated Attendance Report"
# subject for the mail

head4indi = ["Date", "Roll", "Name", "Total Attendance Count",
             "Real", "Duplicate", "Invalid", "Absent"]
# headings for the individual roll attendance file


def setheadindi(wb):
    for i in range(len(head4indi)):
        wb.cell(row=1, column=1+i).value = head4indi[i]    
# function for filling headings 
    


def firstrowset(wb, validunique):
    for i in range(len(validunique)):
        wb.cell(row=3+i, column=1).value = validunique[i]
# function for filling first row of unique dates

def validtime(timetem):
    if timetem == "15:00":
        return True
    hr = timetem[0:2]
    if hr == "14":
        return True
    return False
# function for checking whether the attendance marking time is valid or not

def rollsplit(rollpname):
    ans = ""
    for i in range(8):
        ans += rollpname[i]
    return ans
# splitting roll from Roll Name

def namesplit(rollpname):
    ans = rollpname[9:]
    return ans
# splitting Name from Roll Name

def datesplit(timeofatt):
    ans = timeofatt[0:10]
    return ans
# splitting date from timestamps

def timesplit(timeofatt):
    ans = timeofatt[11:]
    return ans
# splitting time from timestamp

def per(a, b):
    c = a/b
    c *= 100
    return round(c,2)
# function for percentage

start_time = datetime.datetime.now()

registered = pd.read_csv('input_registered_students.csv')

dictofatt = {}
dictofnames = {}
dictoffakes = {}
regroll = []
dictofRoll = {}
# made some dictionaries


os.makedirs('output', exist_ok=True)

cons = openpyxl.Workbook()
consolidated = cons.active

consolidated.cell(row=1, column=1).value = "Roll"
consolidated.cell(row=1, column=2).value = "Name"


for i in range(len(registered)):
    tempr = (str)(registered.iloc[i, 0])
    tempn = (str)(registered.iloc[i, 1])
    dictofnames[tempr] = tempn
    regroll.append(tempr)
    dictofatt[tempr] = []
    rolltemp = openpyxl.Workbook()
    rolltempfile = rolltemp.active
    path = "output/" + (str)(tempr) + ".xlsx"
    rolltemp.save(path)
# made individual roll files

attendance = pd.read_csv('input_attendance.csv')

validdates = []

for i in range(len(attendance)):
    datetemp = datesplit(attendance.iloc[i, 0])
    day = pd.to_datetime(datetemp, format="%d-%m-%Y")
    if (day.day_name() == "Monday" or day.day_name() == "Thursday"):
        # checking whether the day is monday/thursday or not
        validdates.append(datetemp)

validdates = set(validdates)

sortedValidDates = []
for i in validdates:
    sortedValidDates.append(i)

sortedValidDates.sort(key=lambda date: datetime.datetime.strptime(date, "%d-%m-%Y"))
# made a list of only valid dates in sorted manner
# print(sortedValidDates)

totalLectures = (len)(validdates)
# total lectures taken
x = 0
for i in regroll:
    dictofRoll[i] = x+1
    consolidated.cell(row=2+x, column=1).value = i
    consolidated.cell(row=2+x, column=2).value = dictofnames[i]
    x += 1
    # filled the consolidated report with rolls and names

head2 = ["Actual Lecture Taken", "Total Real", "% Attendance"]

for i in range(len(head2)):
    consolidated.cell(row=1, column=3+totalLectures+i).value = head2[i]
# filled the heading of consolidated report


for i in range(len(attendance)):

    rollpname = attendance.iloc[i, 1]
    rollt = rollsplit(rollpname)
    namet = namesplit(rollpname)
    rolltemp = openpyxl.Workbook()
    rolltempfile = rolltemp.active
    setheadindi(rolltempfile)
    firstrowset(rolltempfile, sortedValidDates)
    rolltempfile.cell(row=2, column=2).value = rollt
    rolltempfile.cell(row=2, column=3).value = namet
    path = "output/" + (str)(rollt) + ".xlsx"

    timeofatt = attendance.iloc[i, 0]
    timetemp = timesplit(timeofatt)
    datetemp = datesplit(timeofatt)
    day = pd.to_datetime(datetemp, format="%d-%m-%Y")
    rolltemp.save(path)

    dictofatt[rollt].append(timeofatt)
# filled the individual roll files with names,rolls and made a dictionary which has key as rolls and will store the timestamp values

for i in regroll:
    timestamplist = dictofatt[i]
    timestamplist = sorted(
        timestamplist, key=lambda date: datetime.datetime.strptime(date, "%d-%m-%Y %H:%M"))

    totalmarked = {}
    validmarked = {}
    duplimarked = {}
    # dictionaries to store the total, valid, duplicate attendance marked
    invalidmarked = 0
    for t in validdates:
        totalmarked[t] = 0
        validmarked[t] = 0
        duplimarked[t] = 0

    # print(totalmarked, validmarked, duplimarked)
    for times in timestamplist:
        dateof = datesplit(times)
        timeof = timesplit(times)

        if dateof in validdates:
            totalmarked[dateof] += 1
            if validtime(timeof):
                if validmarked[dateof] == 0:
                    validmarked[dateof] += 1
                else:
                    duplimarked[dateof] += 1
                    # print(totalmarked[dateof], validmarked[dateof], duplimarked[dateof])
        else:
            invalidmarked += 1
    # checking the total valid, and other counts of each roll
    
    path = "output/" + (str)(i) + ".xlsx"
    # filename = (str)(i) + ".xlsx"
    tempwb = load_workbook(path)
    tempsheet = tempwb.active

    # print(duplimarked[dateofroll])
    # print(totalmarked["2001EE83"])

    x = 0
    for j in validdates:
        tempsheet.cell(row=3+x, column=4).value = totalmarked[j]
        tempsheet.cell(row=3+x, column=5).value = validmarked[j]
        tempsheet.cell(row=3+x, column=6).value = duplimarked[j]
        tempsheet.cell(row=3+x, column=7).value = totalmarked[j]-validmarked[j]-duplimarked[j]
        if validmarked[j] > 0:
            tempsheet.cell(row=3+x, column=8).value = "YES"
            consolidated.cell(row=1+dictofRoll[i], column=3+x).value = "P"
        else:
            tempsheet.cell(row=3+x, column=8).value = "NO"
            consolidated.cell(row=1+dictofRoll[i], column=3+x).value = "A"
        x += 1
        # updated the individual, and consolidated reports

    tempwb.save(path)

# print(dictofatt)

for i in range(len(dictofRoll)):
    x = 0
    for j in range(totalLectures):
        val = consolidated.cell(row=2+i, column=3+j).value
        if val == "P":
            x += 1
    consolidated.cell(row=2+i, column=3+totalLectures).value = totalLectures
    consolidated.cell(row=2+i, column=4+totalLectures).value = x
    consolidated.cell(row=2+i, column=5+totalLectures).value = per(x, totalLectures)
# filled the percentage and remaining parts in consolidated report

for i in range(len(sortedValidDates)):
    consolidated.cell(row=1, column=3+i).value = sortedValidDates[i]

cons.save("output/input_attendance_consolidated.xlsx")
# saved the consolidated report

end_time1 = datetime.datetime.now()

print("Reports Created in: {}".format(end_time1-start_time))


res = input("[+] Do you want to send Consolidated Attendance Report to cs3842022@gmail.com? (Y/N)\n")
filepath = "output/input_attendance_consolidated.xlsx"
# asking that whether user wants to send mail or not
msg_body = '''Respected Sir, 

Please find the consolidated attendance report attached in the mail.

Thanks and Regards
Yash Raj
Roll-2001EE83
Electrical & Electronics Engg.
Ph No. +91-7903510948'''

if(res=='Y'):
    send_mail(FROM_ADDR, FROM_PASSWD, TO_ADDR, "Consolidated Attendance Report", msg_body, filepath)
    # sent the mail
    end_time2 = datetime.datetime.now()
    print("Mail Sent in: {}".format(end_time2-end_time1))

end_time3 = datetime.datetime.now()
print("Program executed in: {}".format(end_time3-start_time))
# ----------------------------------------------------------------------------------------------------------------------------
    